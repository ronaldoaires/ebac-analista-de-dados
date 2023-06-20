# Módulo | Análise de dados: Coleta de Dados I

# 2. Excel para JSON - Utilizando o pacote Python openpyxl, extraia os dados das colunas escolaridade e tipo_cartao, removendo duplicados. Com os dados, construa o dicionário Python credito. Para finalizar, converta o dicionário credito no formato JSON.

from openpyxl import load_workbook
import json

# Carrega o arquivo Excel
planilha = load_workbook('./arquivos/10/credito.xlsx')

# Obtém a planilha ativa
folha = planilha.active

# Conjuntos para armazenar valores únicos de escolaridade e tipo de cartão
escolaridade = set()
tipo_cartao = set()

# Percorre as linhas da planilha a partir da segunda linha
for linhas in folha.iter_rows(min_row=2):
    # Extrai os valores das células de escolaridade e tipo de cartão da linha atual
    e, c = [linhas[5].value, linhas[8].value]
    
    # Adiciona os valores aos conjuntos
    escolaridade.add(e)
    tipo_cartao.add(c)
    
# Dicionário para armazenar os dados de crédito
credito = {}
credito['escolaridade'] = list(escolaridade)
credito['tipo_cartao'] = list(tipo_cartao)

# Converte o dicionário em uma representação JSON formatada
credito_json = json.dumps(credito, indent=4)

# Imprime o JSON
print(credito_json)
