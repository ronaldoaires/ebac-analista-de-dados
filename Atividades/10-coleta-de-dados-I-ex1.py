# Módulo | Análise de dados: Coleta de Dados I

# 1. Excel para CSV - Utilizando o pacote Python openpyxl, extraia as colunas id, sexo e idade para dos clientes inadimplentes ( default = 1 ) e solteiros (estado_civil = 'solteiro'). Salves os dados extraídos no arquivo csv credito.csv separado por ; e com cabeçalho id;sexo;idade

from openpyxl import load_workbook
import csv

# Carrega o arquivo Excel
planilha = load_workbook('./arquivos/10/credito.xlsx')

# Obtém a planilha ativa
folha = planilha.active

# Lista para armazenar os dados filtrados
clientes = []

# Percorre as linhas da planilha a partir da segunda linha
for linhas in folha.iter_rows(min_row=2):

    # Extrai os valores das células da linha atual
    id, inadimplente, idade, sexo, estado_civil = [
        linhas[0].value, linhas[1].value, linhas[2].value, linhas[3].value, linhas[6].value
    ]

    # Verifica se o cliente é inadimplente e tem estado civil igual a 'solteiro'
    if inadimplente == 1 and estado_civil == 'solteiro':
        # Adiciona os dados do cliente à lista de clientes filtrados
        clientes.append([id, idade, sexo])

# Abre o arquivo CSV no modo de escrita
with open(file='./arquivos/10/credito.csv', mode='w', newline='', encoding='utf8') as arquivo:
    # Cria um objeto writer do módulo csv
    escrever = csv.writer(arquivo, delimiter=';')

    # Escreve os dados no arquivo CSV
    escrever.writerows([['id', 'idade', 'sexo']] + clientes)

# Exibe uma mensagem indicando que o arquivo CSV foi gerado com sucesso
print('Arquivo CSV gerado com sucesso!')

   